from django.contrib import admin
from django.shortcuts import render, redirect
from django.urls import path
from django import forms
from openpyxl import load_workbook
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from .models import Card, Order, Product, PriceTier

# 自定义 Admin 站点标题
admin.site.site_header = '数字商店管理后台'
admin.site.site_title = '数字商店'
admin.site.index_title = '后台管理'


class PriceTierInline(admin.TabularInline):
    """价格阶梯内联编辑"""
    model = PriceTier
    extra = 1
    fields = ('min_quantity', 'max_quantity', 'unit_price', 'display_order')
    ordering = ['display_order', 'min_quantity']
    verbose_name = '价格阶梯'
    verbose_name_plural = '价格阶梯配置'


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ('name', 'price', 'has_tiered_pricing_display', 'display_order', 'stock_count', 'created_at', 'updated_at')
    list_editable = ('display_order',)
    prepopulated_fields = {'slug': ('name',)}
    search_fields = ('name', 'description')
    list_filter = ('created_at',)
    ordering = ['display_order', '-created_at']
    inlines = [PriceTierInline]

    @admin.display(description='阶梯定价', boolean=True)
    def has_tiered_pricing_display(self, obj):
        """显示是否配置了阶梯价格"""
        return obj.has_tiered_pricing()


class CardAdminForm(forms.ModelForm):
    """卡密管理表单"""
    class Meta:
        model = Card
        fields = '__all__'
        widgets = {
            'content': forms.Textarea(attrs={
                'rows': 10,
                'cols': 80,
                'style': 'font-family: monospace; font-size: 14px;',
                'placeholder': '请输入卡密内容，例如：\n- 激活码：XXXX-XXXX-XXXX-XXXX\n- 下载链接：https://...\n- 账号密码：账号xxx 密码xxx'
            }),
        }
        help_texts = {
            'content': '支持多行文本，建议使用等宽字体以便查看'
        }


class ExcelImportForm(forms.Form):
    """Excel 导入表单"""
    product = forms.ModelChoiceField(
        queryset=Product.objects.all(),
        label='选择商品',
        help_text='选择要导入卡密的商品'
    )
    excel_file = forms.FileField(
        label='Excel 文件',
        help_text='上传包含卡密的 Excel 文件（仅支持 .xlsx 格式）。第一列为卡密内容，从第二行开始读取。'
    )


class ProductStockFilter(admin.SimpleListFilter):
    """按产品库存状态筛选"""
    title = '产品库存状态'
    parameter_name = 'product_stock'

    def lookups(self, request, model_admin):
        return (
            ('in_stock', '有库存产品'),
            ('out_of_stock', '无库存产品'),
        )

    def queryset(self, request, queryset):
        if self.value() == 'in_stock':
            # 筛选有未售出卡密的产品
            products_with_stock = Card.objects.filter(
                status='unsold'
            ).values_list('product_id', flat=True).distinct()
            return queryset.filter(product_id__in=products_with_stock)
        elif self.value() == 'out_of_stock':
            # 筛选没有未售出卡密的产品
            products_with_stock = Card.objects.filter(
                status='unsold'
            ).values_list('product_id', flat=True).distinct()
            return queryset.exclude(product_id__in=products_with_stock)
        return queryset


class OrderStatusFilter(admin.SimpleListFilter):
    """按订单关联状态筛选"""
    title = '订单关联'
    parameter_name = 'order_status'

    def lookups(self, request, model_admin):
        return (
            ('with_order', '已关联订单'),
            ('without_order', '未关联订单'),
        )

    def queryset(self, request, queryset):
        if self.value() == 'with_order':
            return queryset.filter(order__isnull=False)
        elif self.value() == 'without_order':
            return queryset.filter(order__isnull=True)
        return queryset


@admin.action(description='导出选中的卡密（Excel）')
def export_cards_to_excel(modeladmin, request, queryset):
    """批量导出卡密到Excel"""
    # 限制导出数量，防止超时
    if queryset.count() > 10000:
        modeladmin.message_user(
            request,
            '一��最多导出10000条卡密，请使用筛选功能分批导出',
            level='error'
        )
        return

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = '卡密导出'

    # 设置表头
    headers = ['ID', '所属商品', '卡密内容', '状态', '关联订单ID', '买家邮箱', '创建时间']
    ws.append(headers)

    # 设置表头样式
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 写入数据（使用select_related优化查询）
    for card in queryset.select_related('product', 'order'):
        ws.append([
            card.id,
            card.product.name,
            card.content,
            card.get_status_display(),
            card.order.id if card.order else '',
            card.order.email if card.order else '',
            card.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ])

    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 20

    # 创建HTTP响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="cards_export_{timestamp}.xlsx"'

    wb.save(response)
    return response


@admin.action(description='批量设置为已售出')
def mark_as_sold(modeladmin, request, queryset):
    """批量将卡密标记为已售出（不关联订单）"""
    # 只更新未售出的卡密
    unsold_cards = queryset.filter(status='unsold')
    count = unsold_cards.count()

    if count == 0:
        modeladmin.message_user(
            request,
            '选中的卡密中没有未售出的，无需更新',
            level='warning'
        )
        return

    # 批量更新状态
    unsold_cards.update(status='sold')

    modeladmin.message_user(
        request,
        f'成功将{count}个卡密设置为已售出。注意：这些卡密未关联订单，仅标记状态。',
        level='success'
    )


@admin.action(description='批量设置为未售出')
def mark_as_unsold(modeladmin, request, queryset):
    """批量将卡密标记为未售出（安全检查）"""
    # 只更新已售出的卡密
    sold_cards = queryset.filter(status='sold')
    count = sold_cards.count()

    if count == 0:
        modeladmin.message_user(
            request,
            '选中的卡密中没有已售出的，无需更新',
            level='warning'
        )
        return

    # 检查是否有关联订单（防止误操作）
    cards_with_orders = sold_cards.filter(order__isnull=False)
    if cards_with_orders.exists():
        modeladmin.message_user(
            request,
            f'错误：选中的{cards_with_orders.count()}个卡密已关联订单，不能改回未售出状态。请在订单管理中处理。',
            level='error'
        )
        return

    # 批量更新状态
    sold_cards.update(status='unsold')

    modeladmin.message_user(
        request,
        f'成功将{count}个卡密设置为未售出',
        level='success'
    )


@admin.register(Card)
class CardAdmin(admin.ModelAdmin):
    form = CardAdminForm
    list_display = ('id', 'product', 'status', 'short_content', 'order', 'created_at')
    list_filter = ('status', 'product', ProductStockFilter, OrderStatusFilter)
    search_fields = ('content', 'product__name')
    autocomplete_fields = ['order']
    list_editable = ('status',)
    ordering = ['-created_at']
    readonly_fields = ('created_at',)

    # 批量操作
    actions = [export_cards_to_excel, mark_as_sold, mark_as_unsold]

    fieldsets = (
        ('基本信息', {
            'fields': ('product', 'status'),
            'description': '卡密的基本配置信息'
        }),
        ('卡密内容', {
            'fields': ('content',),
            'description': '请输入完整的卡密内容。可以是激活码、下载链接、账号密码等。',
            'classes': ('wide',)
        }),
        ('关联信息', {
            'fields': ('order', 'created_at'),
            'description': '卡密的销售和创建信息'
        }),
    )

    @admin.display(description='卡密内容')
    def short_content(self, obj):
        if len(obj.content) > 30:
            return obj.content[:30] + '...'
        return obj.content

    def get_urls(self):
        """添加自定义 URL"""
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel_view), name='card_import_excel'),
        ]
        return custom_urls + urls

    def import_excel_view(self, request):
        """处理 Excel 导入"""
        if request.method == 'POST':
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                product = form.cleaned_data['product']
                excel_file = request.FILES['excel_file']

                # 验证文件格式
                if not excel_file.name.endswith('.xlsx'):
                    messages.error(request, '只支持 .xlsx 格式的 Excel 文件')
                    return redirect('.')

                try:
                    # 加载 Excel 文件
                    workbook = load_workbook(excel_file, read_only=True)
                    sheet = workbook.active

                    # 读取卡密（从第二行开始，第一列）
                    cards_to_create = []
                    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
                        card_content = row[0]
                        if card_content:  # 跳过空行
                            card_content = str(card_content).strip()
                            if card_content:
                                cards_to_create.append(Card(
                                    product=product,
                                    content=card_content,
                                    status='unsold'
                                ))

                    # 批量创建卡密
                    if cards_to_create:
                        Card.objects.bulk_create(cards_to_create)
                        messages.success(request, f'成功导入 {len(cards_to_create)} 个卡密到商品「{product.name}」')
                    else:
                        messages.warning(request, '未找到有效的卡密数据')

                    return redirect('..')

                except Exception as e:
                    messages.error(request, f'导入失败：{str(e)}')
                    return redirect('.')
        else:
            form = ExcelImportForm()

        context = {
            'form': form,
            'title': 'Excel 批量导入卡密',
            'site_header': admin.site.site_header,
            'site_title': admin.site.site_title,
            'has_permission': True,
        }
        return render(request, 'admin/card_import_excel.html', context)

    # 在列表页添加导入按钮
    change_list_template = 'admin/card_change_list.html'


@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = ('id', 'email', 'quantity', 'total_amount', 'status', 'created_at')
    list_filter = ('status', 'created_at')
    search_fields = ('email', 'out_trade_no')
    readonly_fields = ('created_at',)
    list_editable = ('status',)
    ordering = ['-created_at']
